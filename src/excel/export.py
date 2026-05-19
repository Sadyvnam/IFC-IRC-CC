import pandas as pd
from dataclasses import asdict
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from src.excel.types import Report
from src.wlca_functions.check_metrics import (
    category_context_from_overview,
    check_definition_rows,
    methodology_notes,
    metrics_to_count_rows,
    metrics_to_indicator_rows,
)

def export_excel(report: Report, path: str) -> None:
    overview_df=pd.DataFrame()
    category_context = category_context_from_overview(getattr(report, "overviewTable", []) or [])
    try:
        check_metrics = getattr(report, "check_metrics", []) or []
        if check_metrics:
            overview_df = pd.DataFrame(metrics_to_indicator_rows(report.overview.indicators, check_metrics, category_context))
        else:
            overview_df = pd.DataFrame([r for r in report.overview.indicators])
    except:
        print("Couldnt handle indicator export")
        pass
    #indicators_df = pd.DataFrame([r for r in report.overview.])
    issues_df     = pd.DataFrame([r for r in report['issues']])
    elements_df   = pd.DataFrame([r for r in report['elements']])
    overview_table= pd.DataFrame([r for r in report['overviewTable']])
    check_counts_df = pd.DataFrame()
    try:
        check_metrics = getattr(report, "check_metrics", []) or []
        if check_metrics:
            check_counts_df = pd.DataFrame(metrics_to_count_rows(report.overview.indicators, check_metrics, category_context))
    except:
        print("Couldnt handle check count export")
    methodology_df = pd.DataFrame(methodology_notes())
    check_definitions_df = pd.DataFrame(check_definition_rows())

    for df in [overview_table, overview_df, issues_df, elements_df, check_counts_df, check_definitions_df]:
        num_cols = df.select_dtypes(include="number").columns
        df[num_cols] = df[num_cols].round(3)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overview_table.to_excel(writer, sheet_name="Overview", index=False)
        overview_df.to_excel(writer, sheet_name="Indicators", index=False)
        if not check_counts_df.empty:
            check_counts_df.to_excel(writer, sheet_name="Check Counts", index=False)
        check_definitions_df.to_excel(writer, sheet_name="Check Definitions", index=False)
        methodology_df.to_excel(writer, sheet_name="Methodology Notes", index=False)
        issues_df.to_excel(writer,     sheet_name="Issues",     index=False)
        elements_df.to_excel(writer,   sheet_name="Elements",   index=False)

        wb = writer.book

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                header = cell.value
                if isinstance(header, str) and header.endswith("_pct"):
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=cell.column).number_format = "0.0%"
            ws.move_range
